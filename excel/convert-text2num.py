from openpyxl import load_workbook

def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return None

def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return None

def process_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if type(cell.value) == str:
                if is_int(cell.value) != None:
                    print('cell {0} converted to int {1}'.format(cell, cell.value))
                    cell.value = int(cell.value)
                elif is_float(cell.value) != None:
                    print('cell {0} converted float {1}'.format(cell, cell.value))
                    cell.value = float(cell.value)

def process_workbook(source, dest):
    wb = load_workbook(source)

    for sheet in wb.worksheets:
        process_sheet(sheet)

    wb.save(dest)

if __name__ == "__main__":
    print('please input source xlsx file path:')
    source = input()

    print('please input dest xlsx file path:')
    dest = input()
    process_workbook(source, dest)

    print('\ndone!')
