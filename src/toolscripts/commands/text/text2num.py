"""``text2num`` - convert numeric-looking text in an Excel workbook to real numbers."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from toolscripts.core.log import add_logging_flags, configure_from_args, get_logger

log = get_logger(__name__)


def _is_int(value: str) -> bool:
    try:
        int(value)
    except (TypeError, ValueError):
        return False
    return True


def _is_float(value: str) -> bool:
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True


def _process_workbook(source: Path, dest: Path) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:
        log.error("missing dependency: install with `pip install openpyxl`")
        sys.exit(1)

    wb = load_workbook(source)
    converted = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if _is_int(cell.value):
                        cell.value = int(cell.value)
                        converted += 1
                    elif _is_float(cell.value):
                        cell.value = float(cell.value)
                        converted += 1
    wb.save(dest)
    log.success("converted %d cells -> %s", converted, dest)


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="text2num",
        description="Convert numeric-looking text in an .xlsx workbook to real numbers.",
    )
    parser.add_argument("source", help="source .xlsx file")
    parser.add_argument("dest", help="destination .xlsx file")
    add_logging_flags(parser)
    args = parser.parse_args()
    configure_from_args(args)

    source = Path(args.source).expanduser()
    dest = Path(args.dest).expanduser()
    if not source.is_file():
        log.error("source not found: %s", source)
        sys.exit(1)
    _process_workbook(source, dest)


if __name__ == "__main__":
    main()
